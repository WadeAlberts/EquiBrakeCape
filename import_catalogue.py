# =============================================================================
# import_catalogue.py
# =============================================================================
# Reads the Vehicle Catalogue Excel file (~30,000 rows) and:
# 1. Creates a vehicle record for each row
# 2. For every part number found in that row, creates a fitment record
#    linking that vehicle to that part number
#
# Run AFTER import_prices.py
# Command: python3 import_catalogue.py
#
# BEFORE RUNNING:
# - Place your catalogue Excel file in the same folder as this script
# - Make sure you have already run import_prices.py
# =============================================================================

import sqlite3
import openpyxl
import os

# ------------------------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------------------------

CATALOGUE_FILE = 'Vehicle Catalogue.xlsx'
SHEET_NAME     = 'Applications'
DATABASE       = 'products.db'

# ------------------------------------------------------------------------------
# COLUMN DEFINITIONS
# These are the columns that contain VEHICLE INFO (not part numbers).
# We store them in the vehicles table directly.
# Everything else is treated as a potential part number column.
# ------------------------------------------------------------------------------

# Maps the exact Excel column heading → the database column name in vehicles table
VEHICLE_INFO_COLUMNS = {
    'Make'                    : 'make',
    'Model'                   : 'model',
    'Sub-Model'               : 'sub_model',
    'Engine'                  : 'engine',
    'Engine Type'             : 'engine_type',
    'Valves'                  : 'valves',
    'BHP'                     : 'bhp',
    'Year'                    : 'year',
    'Special Comments'        : 'special_comments',
    'Front Brake Caliper Type': 'front_caliper_type',
    'Rear Brake Caliper Type' : 'rear_caliper_type',
    'Front Brake Shoe Type'   : 'front_shoe_type',
    'Rear Brake Shoe Type'    : 'rear_shoe_type',
}

# These are SPEC columns — they describe disc specs, not part numbers.
# We store them in the vehicles table too but they appear TWICE
# (once for front, once for rear) so we handle them by position.
SPEC_COLUMNS = [
    'Solid or Vented',
    'Number of Bolt Holes',
    'Diameter',
    'Total Height',
    'Thickness New/Min',
]

# These columns contain KIT CONTENTS descriptions, not part numbers.
# We skip them entirely during fitment import.
SKIP_IF_CONTAINS = [
    'Kit contents',
    'Kit Contents',
    'Contents',
]


# ------------------------------------------------------------------------------
# HELPER FUNCTIONS
# ------------------------------------------------------------------------------

def is_skip_column(col_name):
    """
    Returns True if this column should be skipped entirely.
    These are 'kit contents' description columns, not part numbers.
    """
    for skip_word in SKIP_IF_CONTAINS:
        if skip_word in col_name:
            return True
    return False


def get_position(col_name):
    """
    Tries to determine if a part is for the Front or Rear
    based on the column heading name.
    """
    col_upper = col_name.upper()
    if 'FRONT' in col_upper:
        return 'Front'
    elif 'REAR' in col_upper:
        return 'Rear'
    else:
        return ''


def clean_cell(value):
    """
    Cleans a cell value - strips whitespace and returns None if empty.
    """
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned if cleaned and cleaned.lower() != 'none' else None


def import_catalogue():
    """
    Main import function.
    """

    # --- Checks ---
    if not os.path.exists(CATALOGUE_FILE):
        print(f"❌ File not found: {CATALOGUE_FILE}")
        print("   Place the Excel file in the same folder as this script.")
        return

    if not os.path.exists(DATABASE):
        print(f"❌ Database not found.")
        print("   Run: python3 database.py")
        return

    print(f"📂 Opening: {CATALOGUE_FILE}  (this may take a moment for 30,000 rows...)")

    # read_only=True is faster for large files - we are only reading, not writing
    workbook = openpyxl.load_workbook(CATALOGUE_FILE, read_only=True, data_only=True)

    if SHEET_NAME not in workbook.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' not found in the file.")
        print(f"   Sheets available: {workbook.sheetnames}")
        return

    sheet = workbook[SHEET_NAME]
    print(f"✅ Sheet '{SHEET_NAME}' opened.")

    # --- Read the header row to map column names to column index numbers ---
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip() if h else '' for h in row]
    print(f"📋 Found {len(headers)} columns in the catalogue.")

    # --- Connect to database ---
    connection = sqlite3.connect(DATABASE)
    cursor = connection.cursor()

    # Track our progress
    vehicles_imported  = 0
    fitments_imported  = 0
    rows_skipped       = 0

    # We need to track how many times we've seen a spec column
    # because Solid or Vented / Diameter etc appear twice (front + rear)
    # We handle this by tracking the index position of those columns

    # Build a lookup: index → column heading
    # Also identify the two sets of spec columns (front disc vs rear disc)
    front_spec_done = False   # We'll flip this when we see the first set of specs
    col_roles = []            # For each column index: its role

    for i, h in enumerate(headers):
        if h in VEHICLE_INFO_COLUMNS:
            col_roles.append(('vehicle_info', VEHICLE_INFO_COLUMNS[h]))
        elif h in SPEC_COLUMNS:
            if not front_spec_done:
                col_roles.append(('front_spec', h))
                # Once we've mapped all 5 spec columns once, flip the flag
                if h == 'Thickness New/Min':
                    front_spec_done = True
            else:
                col_roles.append(('rear_spec', h))
        elif is_skip_column(h):
            col_roles.append(('skip', h))
        elif h == '':
            col_roles.append(('skip', 'empty'))
        else:
            # Everything else is a potential part number column
            col_roles.append(('part_number', h))

    print(f"🔍 Column roles assigned. Starting data import...\n")

    # --- Process each data row ---
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

        # Skip completely empty rows
        if not any(row):
            rows_skipped += 1
            continue

        # --- Build the vehicle record from vehicle_info columns ---
        vehicle_data = {
            'make': None, 'model': None, 'sub_model': None,
            'engine': None, 'engine_type': None, 'valves': None,
            'bhp': None, 'year': None, 'special_comments': None,
            'front_caliper_type': None, 'rear_caliper_type': None,
            'front_shoe_type': None, 'rear_shoe_type': None,
            'front_solid_vented': None, 'front_bolt_holes': None,
            'front_diameter': None, 'front_total_height': None,
            'front_thickness': None,
            'rear_solid_vented': None, 'rear_bolt_holes': None,
            'rear_diameter': None, 'rear_total_height': None,
            'rear_thickness': None,
        }

        # Map spec column names to vehicle_data keys
        front_spec_map = {
            'Solid or Vented'    : 'front_solid_vented',
            'Number of Bolt Holes': 'front_bolt_holes',
            'Diameter'           : 'front_diameter',
            'Total Height'       : 'front_total_height',
            'Thickness New/Min'  : 'front_thickness',
        }
        rear_spec_map = {
            'Solid or Vented'    : 'rear_solid_vented',
            'Number of Bolt Holes': 'rear_bolt_holes',
            'Diameter'           : 'rear_diameter',
            'Total Height'       : 'rear_total_height',
            'Thickness New/Min'  : 'rear_thickness',
        }

        # Collect part numbers found in this row
        fitments_to_add = []

        for i, cell_value in enumerate(row):
            if i >= len(col_roles):
                break

            role, role_name = col_roles[i]
            value = clean_cell(cell_value)

            if role == 'vehicle_info':
                vehicle_data[role_name] = value

            elif role == 'front_spec' and value:
                db_col = front_spec_map.get(role_name)
                if db_col:
                    vehicle_data[db_col] = value

            elif role == 'rear_spec' and value:
                db_col = rear_spec_map.get(role_name)
                if db_col:
                    vehicle_data[db_col] = value

            elif role == 'part_number' and value:
                # This cell contains a part number
                # role_name is the column heading e.g. "EBC Greenstuff Front Pads"
                position = get_position(role_name)
                fitments_to_add.append({
                    'part_number' : value,
                    'product_type': role_name,
                    'position'    : position,
                })

        # Skip rows with no make (they are empty or header repeats)
        if not vehicle_data['make']:
            rows_skipped += 1
            continue

        # --- Insert the vehicle into the database ---
        cursor.execute('''
            INSERT INTO vehicles (
                make, model, sub_model, engine, engine_type, valves, bhp, year,
                special_comments, front_caliper_type, rear_caliper_type,
                front_shoe_type, rear_shoe_type,
                front_solid_vented, front_bolt_holes, front_diameter,
                front_total_height, front_thickness,
                rear_solid_vented, rear_bolt_holes, rear_diameter,
                rear_total_height, rear_thickness
            ) VALUES (
                :make, :model, :sub_model, :engine, :engine_type, :valves,
                :bhp, :year, :special_comments, :front_caliper_type,
                :rear_caliper_type, :front_shoe_type, :rear_shoe_type,
                :front_solid_vented, :front_bolt_holes, :front_diameter,
                :front_total_height, :front_thickness,
                :rear_solid_vented, :rear_bolt_holes, :rear_diameter,
                :rear_total_height, :rear_thickness
            )
        ''', vehicle_data)

        # Get the ID of the vehicle we just inserted
        vehicle_id = cursor.lastrowid
        vehicles_imported += 1

        # --- Insert a fitment record for each part number found ---
        for fitment in fitments_to_add:
            cursor.execute('''
                INSERT INTO vehicle_fitment (vehicle_id, part_number, product_type, position)
                VALUES (?, ?, ?, ?)
            ''', (
                vehicle_id,
                fitment['part_number'],
                fitment['product_type'],
                fitment['position'],
            ))
            fitments_imported += 1

        # Print progress every 1000 rows so you know it's still running
        if vehicles_imported % 1000 == 0:
            print(f"   ⏳ {vehicles_imported} vehicles processed...")
            connection.commit()  # Save every 1000 rows to avoid memory issues

    # Final save
    connection.commit()
    connection.close()

    print(f"\n🎉 Catalogue import complete!")
    print(f"   🚗 Vehicles imported  : {vehicles_imported}")
    print(f"   🔩 Fitments imported  : {fitments_imported}")
    print(f"   ⏭️  Rows skipped       : {rows_skipped}")
    print(f"\nYou can now run: python3 app.py")


if __name__ == '__main__':
    import_catalogue()