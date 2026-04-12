# =============================================================================
# import_prices.py
# =============================================================================
# Reads the RSA Dealer Pricelist Excel file and imports all part numbers
# and pricing into the products table in our database.
#
# Run BEFORE import_catalogue.py so that part numbers exist before
# fitment records try to reference them.
#
# Command: python3 import_prices.py
#
# BEFORE RUNNING:
# - Place your pricelist Excel file in the same folder as this script
# - Update the filename below if it differs from the default
# =============================================================================

import sqlite3
import openpyxl
import os

# ------------------------------------------------------------------------------
# CONFIGURATION — update these if your file is named differently
# ------------------------------------------------------------------------------

# The name of your pricelist Excel file
PRICE_FILE = 'RSA DEALER PRICELIST AUGUST 2023.xlsx'

# The name of our database
DATABASE = 'products.db'


# ------------------------------------------------------------------------------
# HELPER: Determine product category from part number prefix
# EBC uses consistent prefixes on their part numbers
# ------------------------------------------------------------------------------

def get_category(part_number):
    """
    Guesses the product category based on the EBC part number prefix.
    EBC part numbers follow a pattern:
      DP = Brake Pads
      D   = Brake Disc / Rotor
      GD  = Grooved Disc
      USR = Ultimax Slotted Rotor
      BSD = BSD Grooved Disc
      BLA/BLM = Brake Lines (Automotive/Motorcycle)
      SG  = 2-Piece Floating Disc
      FA  = Motorcycle Brake Pads
      MD  = Motorcycle Disc
      BF  = Brake Fluid
      PD  = Pad & Disc Kit (starts with PD and has K)
    """
    if not part_number:
        return 'Unknown'

    p = str(part_number).strip().upper()

    if p.startswith('DP'):
        return 'Brake Pads'
    elif p.startswith('GD'):
        return 'Grooved Disc'
    elif p.startswith('USR'):
        return 'Ultimax Slotted Rotor'
    elif p.startswith('BSD'):
        return 'BSD Grooved Disc'
    elif p.startswith('SG'):
        return '2-Piece Floating Disc'
    elif p.startswith('D'):
        return 'Brake Disc'
    elif p.startswith('BLA') or p.startswith('BLM') or p.startswith('BLR'):
        return 'Brake Line'
    elif p.startswith('FA') or p.startswith('FD') or p.startswith('MA'):
        return 'Motorcycle Brake Pads'
    elif p.startswith('MD'):
        return 'Motorcycle Disc'
    elif p.startswith('BF'):
        return 'Brake Fluid'
    elif p.startswith('PD') and 'K' in p:
        return 'Brake Kit'
    elif p.startswith('H'):
        return 'Brake Shoe'
    else:
        return 'Other'


def clean_price(value):
    """
    Cleans a price value from Excel.
    Excel sometimes returns prices as strings with symbols or spaces.
    This function converts them to a plain number (float) or None.
    """
    if value is None:
        return None
    try:
        # Remove any currency symbols, spaces, or commas and convert to float
        cleaned = str(value).replace('R', '').replace(',', '').replace(' ', '').strip()
        return float(cleaned) if cleaned else None
    except (ValueError, TypeError):
        return None


def import_prices():
    """
    Main function that reads the Excel pricelist and imports into the database.
    """

    # Check the file exists before doing anything
    if not os.path.exists(PRICE_FILE):
        print(f"❌ File not found: {PRICE_FILE}")
        print("   Make sure the Excel file is in the same folder as this script.")
        return

    if not os.path.exists(DATABASE):
        print(f"❌ Database not found: {DATABASE}")
        print("   Run this first: python3 database.py")
        return

    print(f"📂 Opening: {PRICE_FILE}")

    # Load the Excel workbook
    # data_only=True means we get the calculated values, not the formulas
    workbook = openpyxl.load_workbook(PRICE_FILE, data_only=True)

    print(f"📋 Sheets found: {workbook.sheetnames}")

    # Connect to database
    connection = sqlite3.connect(DATABASE)
    cursor = connection.cursor()

    total_imported = 0
    total_skipped = 0

    # Loop through every sheet in the workbook
    # The pricelist has multiple sheets but they all have the same structure
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"\n   Processing sheet: {sheet_name}")

        sheet_count = 0

        # Loop through every row starting from row 2 (row 1 is the header)
        for row in sheet.iter_rows(min_row=2, values_only=True):

            # The columns in order are:
            # 0: EBC NUMBER
            # 1: DEALER 1 PRICE (less 15%) EXCL VAT
            # 2: DEALER 2 PRICE (less 20%) EXCL VAT
            # 3: DEALER 3 PRICE (less 30%) EXCL VAT
            # 4: SRP EXCL VAT
            # 5: SRP INCL VAT

            # Skip completely empty rows
            if not row[0]:
                continue

            part_number = str(row[0]).strip()

            # Skip rows where the first cell isn't a real part number
            # (sometimes there are section headers or notes in the sheet)
            if not part_number or len(part_number) < 2:
                total_skipped += 1
                continue

            dealer1 = clean_price(row[1] if len(row) > 1 else None)
            dealer2 = clean_price(row[2] if len(row) > 2 else None)
            dealer3 = clean_price(row[3] if len(row) > 3 else None)
            srp_excl = clean_price(row[4] if len(row) > 4 else None)
            srp_incl = clean_price(row[5] if len(row) > 5 else None)

            category = get_category(part_number)

            # INSERT OR REPLACE means:
            # If the part number doesn't exist yet → insert a new row
            # If it already exists → update it with the new values
            cursor.execute('''
                INSERT OR REPLACE INTO products
                    (part_number, category, dealer1_price, dealer2_price,
                     dealer3_price, srp_excl_vat, srp_incl_vat)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (part_number, category, dealer1, dealer2, dealer3, srp_excl, srp_incl))

            total_imported += 1
            sheet_count += 1

        print(f"   ✅ {sheet_count} part numbers imported from {sheet_name}")

    # Save all changes
    connection.commit()
    connection.close()

    print(f"\n🎉 Import complete!")
    print(f"   ✅ Total imported : {total_imported}")
    print(f"   ⏭️  Total skipped  : {total_skipped}")
    print(f"\nYou can now run: python3 import_catalogue.py")


if __name__ == '__main__':
    import_prices()